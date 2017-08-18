# -*- coding: UTF-8 -*-
import sys
import string
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import pprint
import mysql.connector
from datetime import datetime, date

# 字段名称
ROW_GIFT_ID = 'gift_id'
PRD_CODE = 'prd_code'
PRD_NAME = 'prd_name'
PRD_POINT = 'origin_point'
PRD_TYPE = 'type'
PRD_EX_TIMES = 'ex_times'
PRD_MONTH_EXCHANGE = 'month_exchange'
PRD_START_TIME = 'start_time'
PRD_END_TIME = 'end_time'
PRD_EX_TYPE = 'exchange_type'
PRD_STOCK_COUNT = 'total_count'
PRD_MODULE = 'module'
PRD_SEQ_NO = 'seq_no'
PRD_STORE = 'store_scope'
PRD_CARD = 'card_id'
ROW_IS_DEL = 'is_del'
PRD_DESC = 'description'

# xlsx字母列对应的字段名
PRD_COLUMN_NAME_DICT = {}
PRD_COLUMN_NAME_DICT['D'] = ROW_GIFT_ID
PRD_COLUMN_NAME_DICT['E'] = PRD_CODE
PRD_COLUMN_NAME_DICT['F'] = PRD_NAME
PRD_COLUMN_NAME_DICT['G'] = PRD_POINT
PRD_COLUMN_NAME_DICT['I'] = PRD_TYPE
PRD_COLUMN_NAME_DICT['J'] = PRD_EX_TIMES
PRD_COLUMN_NAME_DICT['K'] = PRD_MONTH_EXCHANGE
PRD_COLUMN_NAME_DICT['L'] = PRD_START_TIME
PRD_COLUMN_NAME_DICT['M'] = PRD_END_TIME
PRD_COLUMN_NAME_DICT['N'] = PRD_EX_TYPE
PRD_COLUMN_NAME_DICT['O'] = PRD_STOCK_COUNT
PRD_COLUMN_NAME_DICT['Q'] = PRD_MODULE
PRD_COLUMN_NAME_DICT['R'] = PRD_SEQ_NO
PRD_COLUMN_NAME_DICT['S'] = PRD_STORE
PRD_COLUMN_NAME_DICT['T'] = PRD_CARD
PRD_COLUMN_NAME_DICT['U'] = ROW_IS_DEL
PRD_COLUMN_NAME_DICT['V'] = PRD_DESC

# 用礼品的条形码构建其描述内容
def buildGiftDesc(code):
	url1 = r'<div align="center"><img src="http://cdn-yyj.4000916916.com/wx/wxExchange/prdDescImage/'
	url2 = "{}_{}".format(code, date.today().strftime("%Y%m%d"))
	url3 = r'.jpg"></div>'
	return (url1+url2+url3)

def strippedString(str):
	return string.strip(str)

def formatRowDict(dict):
	# 礼品编码去掉空格
	dict[ROW_GIFT_ID] = strippedString(dict[ROW_GIFT_ID])
	# 产品条形码去掉空格
	dict[PRD_CODE] = strippedString(dict[PRD_CODE])
	# 礼品名称去掉空格
	dict[PRD_NAME] = strippedString(dict[PRD_NAME])
	# 常规兑分值转化为整数
	# dict[PRD_POINT] = int(dict[PRD_POINT])
	# 产品类型截取整数
	dict[PRD_TYPE] = int(dict[PRD_TYPE][0])
	# 限购次数转化为整数
	# dict[PRD_EX_TIMES] = int(dict[PRD_EX_TIMES])
	# 是否每月限制截取整数
	dict[PRD_MONTH_EXCHANGE] = int(dict[PRD_MONTH_EXCHANGE][0])
	# 是否快递截取整数
	dict[PRD_EX_TYPE] = int(dict[PRD_EX_TYPE][0])
	# 所在版块截取整数
	dict[PRD_MODULE] = int(dict[PRD_MODULE][0])
	# 所在位置转化为整数
	# dict[PRD_SEQ_NO] = int(dict[PRD_SEQ_NO])
	# 是否指定门店截取整数
	dict[PRD_STORE] = int(dict[PRD_STORE][0])
	# 是否券码截取整数
	dict[PRD_CARD] = int(dict[PRD_CARD][0])
	# 是否上架截取整数
	dict[ROW_IS_DEL] = int(dict[ROW_IS_DEL][0])
	dict[PRD_DESC] = buildGiftDesc(dict[PRD_CODE])
	return dict

# 将某一行的数据构建成一个字典
def buildRowDict(worksheet, row_index):
	row_dict = {}
	for row in worksheet.iter_rows(min_row=row_index, max_row=row_index):
		for cell in row:
			if cell.column in PRD_COLUMN_NAME_DICT:
				row_dict.setdefault(PRD_COLUMN_NAME_DICT[cell.column], cell.value)
	return formatRowDict(row_dict)

# 数据库查询在上架状态的礼品记录
def selectGiftRow(cursor, gift_xls):
	select_stmt = "SELECT * FROM product WHERE gift_id = %(gift_id)s AND is_del = 0 AND module = %(module)s"
	cursor.execute(select_stmt, {'gift_id': gift_xls[ROW_GIFT_ID], 'module': gift_xls[PRD_MODULE]})
	gift_row = cursor.fetchone()
	return gift_row

# 插入一条新的礼品记录
def insertGiftRow(cursor, gift_xls):
	# 如果Excel表写的是下架，则不要插入新的记录。
	if gift_xls[ROW_IS_DEL] == 0:
		# 插入礼品
		insert_gift = ("INSERT INTO product "
	                   "(gift_id, prd_code, prd_name, origin_point, type, ex_times, start_time, end_time, exchange_type, seq_no, module, description, store_scope, card_id, is_del, month_exchange) "
	                   "VALUES (%(gift_id)s, %(prd_code)s, %(prd_name)s, %(origin_point)s, %(type)s, %(ex_times)s, %(start_time)s, %(end_time)s, %(exchange_type)s, %(seq_no)s, %(module)s, %(description)s, %(store_scope)s, %(card_id)s, %(is_del)s, %(month_exchange)s)")
		cursor.execute(insert_gift, gift_xls)

		# 修改礼物记录里的库存ID
		update_gift = "UPDATE product SET stock_id = %(stock_id)s WHERE id = %(id)s"
		cursor.execute(update_gift, {'stock_id': cursor.lastrowid, 'id': cursor.lastrowid})

		# 插入一条新的库存记录
		insertStock(cursor, gift_xls[PRD_STOCK_COUNT])

# 修改礼品记录为下架状态
def unshelveGiftRow(cursor, id):
	update_gift = "UPDATE product SET is_del = %(is_del)s WHERE id = %(id)s"
	cursor.execute(update_gift, {'is_del': 1, 'id': id})

# 插入一条新的库存记录
def insertStock(cur, total):
	insert_stock = ("INSERT INTO product_stock (total_count) VALUES (%(total_count)s)")
	cur.execute(insert_stock, {'total_count': total})

# 增加库存
def addStock(conn, cur, stock_id, quantity):
	# 查询原来库存总量
	select_stmt = "SELECT total_count FROM product_stock WHERE id = %(stock_id)s"
	cur.execute(select_stmt, {'stock_id': stock_id})
	if cur.rowcount > 0:
		# 增加库存后的总量
		total = cur.fetchone()['total_count'] + quantity
		print total
		# 修改库存总量
		update_stmt = "UPDATE product_stock SET total_count = %(total_count)s WHERE id = %(stock_id)s"
		cur.execute(update_stmt, {'total_count': total, 'stock_id': stock_id})




wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.get_sheet_by_name(unicode('上架0815', "utf-8"))
	
cnx = mysql.connector.connect(user='root', password='tcbj',
                              host='192.168.103.107',
                              database='wxexchange')
if cnx.is_connected():
	print '---- 已连接数据库 ----'
else:
	print '---- 数据库连接失败 ----'

cursor = cnx.cursor(buffered=True, dictionary=True)

gift_xls = buildRowDict(sheet, 3)
gift_xls[ROW_GIFT_ID] = 'test-005'
gift_xls[ROW_IS_DEL] = 1

# 查询数据库礼品记录
gift_row = selectGiftRow(cursor, gift_xls)
print gift_row

if gift_row: # 数据库礼品记录已存在则根据Excel表修改相关数据
	print "** 礼品已经存在，修改相关数据", gift_xls[ROW_GIFT_ID]

	if gift_row[PRD_MONTH_EXCHANGE] == 1: # 每月兑换有限制
		print "** ** 每月兑换有限制，修改相关数据"
	else: # 每月兑换无限制
		print "** ** 每月兑换无限制"
		if gift_row[ROW_IS_DEL] == 1: # 目前是下架状态
			print "** ** ** 目前是下架状态，直接插入一条新的礼品记录", gift_xls[ROW_GIFT_ID]
		else: # 目前是上架状态
			print "**** 目前是上架状态，先修改为下架状态，再插入一条新的记录", gift_xls[ROW_GIFT_ID]
			unshelveGiftRow(cursor, gift_row['id'])
		# 插入一条新记录
		insertGiftRow(cursor, gift_xls)

else: # 数据库礼品记录不存在则插入一条新记录
	print "-- 礼品不存在，插入一条新记录", gift_xls[ROW_GIFT_ID]
	insertGiftRow(cursor, gift_xls)

cnx.commit()
cursor.close()
cnx.close()